"""
Build an AR aging Excel workbook from an invoice export.

Mirrors the bucketing logic in ar-collections-automation
(1_Trailwise_Toolkit/ar-collections-automation/scripts/ar_collections.py)
so the two reports agree to the penny.

Workbook structure:
  - "Aging Summary" sheet: bucket labels, amounts, total outstanding, a
    BarChart of the four buckets.
  - "Invoice Detail" sheet: one row per open invoice (id, client, due
    date, days past due, balance, bucket).

Usage:
    python3 scripts/ar_aging_excel.py \\
        fixtures/input/invoices.csv \\
        --as-of 2026-06-30 \\
        --out ar_aging_2026-06-30.xlsx
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class Invoice:
    id: str
    client_id: str
    client_name: str
    invoice_date: date
    due_date: date
    amount: float
    amount_paid: float
    status: str

    @property
    def balance_due(self) -> float:
        return self.amount - self.amount_paid

    def days_past_due(self, as_of: date) -> int:
        if self.status == "paid":
            return 0
        return max(0, (as_of - self.due_date).days)

    def bucket(self, as_of: date) -> str:
        if self.status == "paid":
            return "paid"
        dpd = self.days_past_due(as_of)
        if dpd <= 30:
            return "current"
        elif dpd <= 60:
            return "31_60"
        elif dpd <= 90:
            return "61_90"
        else:
            return "90_plus"


# ---------------------------------------------------------------------------
# IO
# ---------------------------------------------------------------------------

def _parse_date(s: str) -> date:
    return datetime.strptime(s.strip(), "%Y-%m-%d").date()


def _parse_amount(s: str) -> float:
    return float(s.strip() or "0")


def load_invoices(csv_path: Path) -> List[Invoice]:
    invoices: List[Invoice] = []
    with csv_path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            invoices.append(
                Invoice(
                    id=row["id"].strip(),
                    client_id=row["client_id"].strip(),
                    client_name=row["client_name"].strip(),
                    invoice_date=_parse_date(row["invoice_date"]),
                    due_date=_parse_date(row["due_date"]),
                    amount=_parse_amount(row["amount"]),
                    amount_paid=_parse_amount(row["amount_paid"]),
                    status=row["status"].strip(),
                )
            )
    return invoices


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

BUCKET_ORDER = ["current", "31_60", "61_90", "90_plus"]


def aggregate(invoices: List[Invoice], as_of: date) -> Dict[str, object]:
    bucket_totals: Dict[str, float] = {b: 0.0 for b in BUCKET_ORDER}
    detail: List[Dict[str, object]] = []
    total_outstanding = 0.0
    for inv in invoices:
        if inv.status == "paid":
            continue
        bal = inv.balance_due
        bucket = inv.bucket(as_of)
        bucket_totals[bucket] += bal
        total_outstanding += bal
        detail.append(
            {
                "id": inv.id,
                "client": inv.client_name,
                "due_date": inv.due_date,
                "days_past_due": inv.days_past_due(as_of),
                "balance": round(bal, 2),
                "bucket": bucket,
            }
        )
    return {
        "as_of": as_of,
        "bucket_totals": {b: round(v, 2) for b, v in bucket_totals.items()},
        "total_outstanding": round(total_outstanding, 2),
        "detail": detail,
    }


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def _import_openpyxl():
    import openpyxl  # type: ignore
    from openpyxl.chart import BarChart, Reference  # type: ignore
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    return openpyxl, BarChart, Reference, Alignment, Font, PatternFill


def write_workbook(agg: Dict[str, object], out_path: Path) -> None:
    openpyxl, BarChart, Reference, Alignment, Font, _ = _import_openpyxl()
    wb = openpyxl.Workbook()

    # --- Aging Summary sheet ---
    ws = wb.active
    ws.title = "Aging Summary"

    as_of = agg["as_of"]
    ws["A1"] = "AR Aging Report"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"As of: {as_of.isoformat()}"
    ws["A2"].font = Font(italic=True)

    headers = ["Bucket", "Days Past Due", "Amount", "% of Outstanding"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    labels = {
        "current": "Current (0-30)",
        "31_60": "31-60 days",
        "61_90": "61-90 days",
        "90_plus": "90+ days",
    }
    totals = agg["bucket_totals"]
    total_outstanding = agg["total_outstanding"]
    row = 5
    for b in BUCKET_ORDER:
        amount = totals.get(b, 0.0)
        ws.cell(row=row, column=1, value=labels[b])
        lo, hi = _bucket_range(b)
        ws.cell(row=row, column=2, value=f"{lo}-{hi}" if hi else f"{lo}+")
        ws.cell(row=row, column=3, value=round(amount, 2))
        ws.cell(row=row, column=3).number_format = "$#,##0.00"
        pct = (amount / total_outstanding) if total_outstanding else 0.0
        ws.cell(row=row, column=4, value=pct)
        ws.cell(row=row, column=4).number_format = "0.00%"
        row += 1
    # Total row
    ws.cell(row=row, column=1, value="Total Outstanding").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(total_outstanding, 2)).number_format = "$#,##0.00"
    ws.cell(row=row, column=3).font = Font(bold=True)

    # Bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.title = "AR Aging by Bucket"
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Bucket"
    data_ref = Reference(
        ws,
        min_col=3, min_row=4,
        max_col=3, max_row=row - 1,
    )
    cats_ref = Reference(
        ws,
        min_col=1, min_row=5,
        max_col=1, max_row=row - 1,
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 16
    chart.height = 8
    ws.add_chart(chart, "F4")

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    # --- Invoice Detail sheet ---
    detail_ws = wb.create_sheet("Invoice Detail")
    detail_headers = ["Invoice ID", "Client", "Due Date", "Days Past Due", "Balance", "Bucket"]
    for col, h in enumerate(detail_headers, start=1):
        c = detail_ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
    for r, row_data in enumerate(agg["detail"], start=2):
        detail_ws.cell(row=r, column=1, value=row_data["id"])
        detail_ws.cell(row=r, column=2, value=row_data["client"])
        detail_ws.cell(row=r, column=3, value=row_data["due_date"])
        detail_ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        detail_ws.cell(row=r, column=4, value=row_data["days_past_due"])
        detail_ws.cell(row=r, column=5, value=row_data["balance"])
        detail_ws.cell(row=r, column=5).number_format = "$#,##0.00"
        detail_ws.cell(row=r, column=6, value=row_data["bucket"])

    for col_letter, width in zip("ABCDEF", [14, 24, 12, 14, 14, 12]):
        detail_ws.column_dimensions[col_letter].width = width

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _bucket_range(b: str):
    if b == "current":
        return (0, 30)
    if b == "31_60":
        return (31, 60)
    if b == "61_90":
        return (61, 90)
    return (91, None)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("input", help="Path to invoices.csv")
    parser.add_argument("--as-of", required=True, help="As-of date (YYYY-MM-DD)")
    parser.add_argument("--out", required=True, help="Output .xlsx path")
    parser.add_argument("--json", default=None, help="Optional JSON report path")
    args = parser.parse_args(argv)

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"error: input not found: {in_path}", file=sys.stderr)
        return 2

    as_of = _parse_date(args.as_of)
    invoices = load_invoices(in_path)
    agg = aggregate(invoices, as_of)
    out_path = Path(args.out)
    write_workbook(agg, out_path)

    if args.json:
        import json
        json_path = Path(args.json)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(
            json.dumps(
                {
                    "as_of": agg["as_of"].isoformat(),
                    "bucket_totals": agg["bucket_totals"],
                    "total_outstanding": agg["total_outstanding"],
                    "open_invoice_count": len(agg["detail"]),
                },
                indent=2,
            ),
            encoding="utf-8",
        )

    print(f"AR aging workbook written: {out_path}")
    print(f"  as_of             = {as_of.isoformat()}")
    print(f"  total_outstanding = ${agg['total_outstanding']:,.2f}")
    for b in BUCKET_ORDER:
        print(f"  {b:>8}           = ${agg['bucket_totals'][b]:,.2f}")
    print(f"  open_invoice_count= {len(agg['detail'])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
